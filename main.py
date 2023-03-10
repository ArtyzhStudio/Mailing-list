from distutils.log import error
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from os.path import basename, curdir

import json

import smtplib
import openpyxl

from openpyxl.styles import PatternFill

import tkinter as tk
import tkinter.ttk as ttk

from tkinter import filedialog

def selectTable():
    global db
    name = filedialog.askopenfilename(initialdir=curdir, initialfile=config['table'], filetypes=(("Microsoft Excel files", "*.xl*"), ("All files", "*.*")))
    if name != '':
        db.delete(0, 'end')
        db.insert(0, name)

def refresh():
    status['text'] = ""
    passed['text'] = ""
    errors['text'] = ""
    config['table'] = db.get()
    config['sheet'] = ds.get()
    config['login'] = log.get()
    config['server'] = serv.get()
    config['number'] = number.get()
    config['start'] = strt.get()
    config['a'] = int(adrs.get())
    config['b'] = int(subject.get())
    config['c'] = int(text.get())
    config['d'] = int(attachment.get())

def save():
    refresh()
    with open("settings.json", "w") as fil:
        fil.write(json.dumps(config))

def logIn():
    refresh()
    try:
        global server
        server = smtplib.SMTP_SSL(config['server'])
        server.set_debuglevel(1)
        server.login(config['login'], pswd.get())
    except:
        errors['text'] = "Доступ к почтовому ящику не получен"
        raise Exception('log in failed')

def mail():
    try:
        logIn()
    except:
        return
    a = config['a']
    b = config['b']
    c = config['c']
    d = config['d']
    try:
        wb = openpyxl.load_workbook(filename=config['table'], read_only=False)
    except:
        errors['text'] = "Невозможно получить доступ к таблице"
        win.update()
        return
    try:
        sheet = wb[config['sheet']]
    except:
        errors['text'] = "Невозможно получить доступ к листу таблицы"
        win.update()
        return
    n = int(number.get())
    m = sheet.max_row
    i = int(strt.get())
    j = 1
    cntOK = 0
    cntFAIL = 0
    status['text'] = "Отправка началась . . ."
    win.update()
    while j <= n and i <= m:
        if sheet.cell(i, a).value != None:
            try:
                to = sheet.cell(i, a).value
                head = sheet.cell(i, b).value
                inner = sheet.cell(i, c).value
                files = sheet.cell(i, d).value
                msg = MIMEMultipart()
                msg['From'] = config['login']
                msg['To'] = to
                msg['Subject'] = head
                msg.attach(MIMEText(inner))
                for t in files.split():
                    with open(t, 'rb') as file:
                        add = MIMEApplication(file.read(), Name = basename(t))
                    add['Content-Disposition'] = 'attachment; filename="%s"' % basename(t)
                    msg.attach(add)
                server.sendmail(config['login'], to, msg.as_string())
            except:
                sheet.cell(i, a).fill = PatternFill("solid", start_color='FF9999')
                sheet.cell(i, b).fill = PatternFill("solid", start_color='FF9999')
                sheet.cell(i, c).fill = PatternFill("solid", start_color='FF9999')
                sheet.cell(i, d).fill = PatternFill("solid", start_color='FF9999')
                cntFAIL += 1
                errors['text'] = "Писем, отправленных с ошибкой : %i" % cntFAIL
                win.update()
            else:
                sheet.cell(i, a).fill = PatternFill("solid", start_color='99FF99')
                sheet.cell(i, b).fill = PatternFill("solid", start_color='99FF99')
                sheet.cell(i, c).fill = PatternFill("solid", start_color='99FF99')
                sheet.cell(i, d).fill = PatternFill("solid", start_color='99FF99')
                cntOK += 1
                passed['text'] = "Писем отосланно : %i" % cntOK
                win.update()
            j += 1
        i += 1
    status['text'] = "Отправка закончилась"
    wb.save(db.get())
    wb.close()
    server.quit()

win = tk.Tk()
win.title('Рассылка писем')
win.geometry('470x500')

try:
    with open("settings.json", "r") as fil:
        config = json.loads(fil.read())
except:
    config = {
	"table": "",
	"sheet": "",
	"login": "",
	"server": "",
    "number" : 0,
    "start" : 1,
	"a": 0,
	"b": 0,
	"c": 0,
	"d": 0
}

database = ttk.Frame(win)
ttk.Label(database, text='Путь к таблице : ').pack(side='left')
ttk.Button(database, text='...', command=selectTable, width=3).pack(side='right')
db = ttk.Entry(database)
db.insert(0, config['table'])
db.pack(fill='x')
database.pack(padx=10, fill='x')

datasheet = ttk.Frame(win)
ttk.Label(datasheet, text='Название листа таблицы : ').pack(side='left')
ds = ttk.Entry(datasheet)
ds.insert(0, config['sheet'])
ds.pack(fill='x')
datasheet.pack(padx=10, fill='x')

ttk.Separator(win, orient='horizontal').pack(padx=36, pady=20, fill='x')

login = ttk.Frame(win)
ttk.Label(login, text='Логин почтового ящика : ').pack(side='left')
log = ttk.Entry(login)
log.insert(0, config['login'])
log.pack(fill='x')
login.pack(padx=10, fill='x')

password = ttk.Frame(win)
ttk.Label(password, text='Пароль приложения почтового ящика : ').pack(side='left')
pswd = ttk.Entry(password, show='\u2022')
pswd.pack(fill='x')
password.pack(padx=10, fill='x')

serversmtp = ttk.Frame(win)
ttk.Label(serversmtp, text='SMTP сервер : ').pack(side='left')
serv = ttk.Entry(serversmtp)
serv.insert(0, config['server'])
serv.pack(fill='x')
serversmtp.pack(padx=10, fill='x')

ttk.Separator(win, orient='horizontal').pack(padx=36, pady=20, fill='x')

start = ttk.Frame(win)
ttk.Label(start, text='Начальная строка писем : ').pack(side='left')
strt = ttk.Entry(start)
strt.insert(0, config['start'])
strt.pack(fill='x')
start.pack(padx=10, fill='x')

num = ttk.Frame(win)
ttk.Label(num, text='Количество отправлямых писем : ').pack(side='left')
number = ttk.Entry(num)
number.insert(0, '5')
number.pack(fill='x')
num.pack(padx=10, fill='x')

ttk.Separator(win, orient='horizontal').pack(padx=36, pady=20, fill='x')

adress = ttk.Frame(win)
ttk.Label(adress, text='Колонка адресов : ').pack(side='left')
adrs = ttk.Entry(adress)
adrs.insert(0, config['a'])
adrs.pack(fill='x')
adress.pack(padx=10, fill='x')

sub = ttk.Frame(win)
ttk.Label(sub, text='Колонка тем писем : ').pack(side='left')
subject = ttk.Entry(sub)
subject.insert(0, config['b'])
subject.pack(fill='x')
sub.pack(padx=10, fill='x')

inn = ttk.Frame(win)
ttk.Label(inn, text='Колонка содержания писем : ').pack(side='left')
text = ttk.Entry(inn)
text.insert(0, config['c'])
text.pack(fill='x')
inn.pack(padx=10, fill='x')

attach = ttk.Frame(win)
ttk.Label(attach, text='Колонка вложений : ').pack(side='left')
attachment = ttk.Entry(attach)
attachment.insert(0, config['d'])
attachment.pack(fill='x')
attach.pack(padx=10, fill='x')

ttk.Separator(win, orient='horizontal').pack(padx=36, pady=20, fill='x')

buttons = ttk.Frame(win)
submit = ttk.Button(buttons, text='Отослать письма', command=mail)
submit.pack(side='left')
ttk.Button(buttons, text='Сохранить настройки', command=save).pack(side='right')
buttons.pack(padx=10, fill='x')

status = ttk.Label(win, text='')
status.pack(padx=10, fill='x')

passed = ttk.Label(win, text='', foreground='green')
passed.pack(padx=10, fill='x')

errors = ttk.Label(win, text='', foreground='red')
errors.pack(padx=10, fill='x')

server = smtplib.SMTP_SSL(config['server'])
win.mainloop()